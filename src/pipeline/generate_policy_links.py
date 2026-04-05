import os
import sys
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
import dashscope
from dashscope import Generation
import uuid
from openpyxl import load_workbook
import requests
from urllib.parse import urlparse

# ==============================
# 自动设置项目根目录
# ==============================
def setup_project_root():
    script_path = os.path.abspath(sys.argv[0] if __name__ == "__main__" else __file__)
    current_dir = os.path.dirname(script_path)
    for _ in range(10):
        if (os.path.isdir(os.path.join(current_dir, 'data')) and 
            os.path.isdir(os.path.join(current_dir, 'src'))):
            os.chdir(current_dir)
            return
        parent = os.path.dirname(current_dir)
        if parent == current_dir:
            break
        current_dir = parent
    raise RuntimeError("无法定位项目根目录！请确保脚本位于 global_esg_risk_map/ 目录结构中。")

setup_project_root()

# ==============================
# 配置
# ==============================
load_dotenv()
dashscope.api_key = os.getenv("DASHSCOPE_API_KEY")

BASIC_DATA_PATH = "data/input/basic_data.xlsx"
OUTPUT_PATH = "data/intermediate/policy_link.xlsx"

# 确保输出目录存在
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# ==============================
# 辅助函数
# ==============================
def load_countries_and_topics():
    """从 basic_data.xlsx 加载国家和 ESG 议题"""
    countries_df = pd.read_excel(BASIC_DATA_PATH, sheet_name="country_list")
    topics_df = pd.read_excel(BASIC_DATA_PATH, sheet_name="topic_list")
    
    # 提取国家信息（使用列名访问）
    countries = []
    for _, row in countries_df.iterrows():
        country_cn = str(row["country_cn"]).strip()
        country_en = str(row["country_en"]).strip()
        iso_alpha3 = str(row["iso_alpha3"]).strip()
        if country_cn and country_en and iso_alpha3 and country_cn.lower() != 'nan':
            countries.append((country_cn, country_en, iso_alpha3))
    
    topics = topics_df.iloc[:, 0].dropna().astype(str).tolist()
    
    print(f"🌍 将处理 {len(countries)} 个国家（来自 basic_data.xlsx 的 country_list）")
    print(f"📌 共 {len(topics)} 个 ESG 议题")
    return countries, topics

def verify_url(url, timeout=5):
    """
    验证 URL 是否可访问（有效性检查）
    返回: (is_valid: bool, status_code: int or None, error_message: str or None)
    """
    try:
        if not url or not url.startswith("http"):
            return False, None, "URL 格式无效"
        
        # 发送 HEAD 请求（更快）
        response = requests.head(
            url,
            timeout=timeout,
            allow_redirects=True,
            headers={'User-Agent': 'ESG-PolicyBot/1.0'}
        )
        
        # 检查状态码（200-299 视为成功）
        if 200 <= response.status_code < 300:
            return True, response.status_code, None
        elif 300 <= response.status_code < 400:
            # 重定向，尝试 GET 请求验证最终目标
            try:
                get_response = requests.get(
                    url,
                    timeout=timeout,
                    allow_redirects=True,
                    headers={'User-Agent': 'ESG-PolicyBot/1.0'}
                )
                if 200 <= get_response.status_code < 300:
                    return True, get_response.status_code, None
            except:
                pass
            return False, response.status_code, f"重定向链接可能无效 (HTTP {response.status_code})"
        else:
            return False, response.status_code, f"HTTP {response.status_code} - 链接无法访问"
            
    except requests.Timeout:
        return False, None, "请求超时 (>5秒)"
    except requests.ConnectionError:
        return False, None, "连接失败 - 网络问题"
    except Exception as e:
        return False, None, f"验证失败: {str(e)[:50]}"

def generate_policy_prompt(country_en, topic):
    return f"""You are a global ESG policy expert. Your task is to find ONE real, official policy or regulation.

Country: {country_en}
Topic: {topic}

CRITICAL REQUIREMENTS:
1. The policy MUST be real and verified to exist
2. The URL MUST be a direct link to an official government or regulatory document
3. The URL must be publicly accessible and open (no login required for viewing)
4. Prefer official policy documents from government websites or official channels
5. If you are not 100% certain the URL is valid and accessible, respond: No valid policy|N/A

Return format: Policy Name|Official URL

Examples of VALID responses:
- EU Corporate Sustainability Reporting Directive|https://ec.europa.eu/info/law/corporate-sustainability-reporting_en
- China Carbon Neutrality Target|https://www.gov.cn/guowuyuan/guowuyuanhuiyi/2021-10/26/content_5644613.htm

Examples of INVALID responses (return No valid policy|N/A for these):
- Made-up policies
- Broken or inaccessible URLs
- URLs requiring login
- Shortened URLs or unclear links

Policy Name|URL:
"""

def parse_llm_response(response_text):
    """
    解析大模型返回的内容，并验证 URL 有效性
    返回: (policy_name, link) 或 (None, None)
    """
    if not response_text or "No valid policy" in response_text or "N/A" in response_text:
        return None, None
    
    parts = response_text.strip().split("|", 1)
    if len(parts) == 2:
        name, link = parts[0].strip(), parts[1].strip()
        
        # 基础格式检查
        if not name or not link or not link.startswith("http"):
            return None, None
        
        # ✅ 关键：验证 URL 是否真的可访问
        is_valid, status_code, error_msg = verify_url(link)
        
        if is_valid:
            return name, link
        else:
            return None, None
    
    return None, None

def select_llm_model():
    """让用户选择大模型和配置参数"""
    print("\n🤖 大模型配置")
    print("="*50)
    print("可用模型：")
    print("1. qwen-turbo (推荐，更快更便宜)")
    print("2. qwen-plus (平衡性能)")
    print("3. qwen-max (最强模型，更准确但更贵)")
    print("\n")
    
    choice = input("请选择模型 (1-3，默认1): ").strip()
    model_map = {"1": "qwen-turbo", "2": "qwen-plus", "3": "qwen-max"}
    model = model_map.get(choice, "qwen-turbo")
    
    print(f"\n✅ 已选择模型: {model}")
    
    # 配置参数
    print("\n⚙️  配置调用参数")
    print("="*50)
    
    try:
        temp_input = input("温度参数 (0.0-1.0，默认0.3): ").strip()
        temperature = float(temp_input) if temp_input else 0.3
        temperature = max(0.0, min(1.0, temperature))
    except ValueError:
        temperature = 0.3
    
    try:
        top_p_input = input("TopP参数 (0.0-1.0，默认0.8): ").strip()
        top_p = float(top_p_input) if top_p_input else 0.8
        top_p = max(0.0, min(1.0, top_p))
    except ValueError:
        top_p = 0.8
    
    try:
        max_tokens_input = input("最大生成长度 (默认500): ").strip()
        max_tokens = int(max_tokens_input) if max_tokens_input else 500
        max_tokens = max(50, min(2000, max_tokens))
    except ValueError:
        max_tokens = 500
    
    retries_input = input("重试次数 (默认3): ").strip()
    max_retries = int(retries_input) if retries_input else 3
    max_retries = max(1, min(10, max_retries))
    
    config = {
        "model": model,
        "temperature": temperature,
        "top_p": top_p,
        "max_tokens": max_tokens,
        "max_retries": max_retries
    }
    
    print(f"\n✅ 模型配置完成:")
    print(f"   模型: {config['model']}")
    print(f"   温度: {config['temperature']}")
    print(f"   TopP: {config['top_p']}")
    print(f"   最大长度: {config['max_tokens']}")
    print(f"   重试次数: {config['max_retries']}")
    print()
    
    return config

def call_qwen(prompt, model_config, country_en=None, topic=None):
    """
    调用大模型，使用用户配置的参数
    如果返回的链接无效，会自动重试获取有效链接
    """
    model = model_config["model"]
    temperature = model_config["temperature"]
    top_p = model_config["top_p"]
    max_tokens = model_config["max_tokens"]
    max_retries = model_config["max_retries"]
    
    valid_link_attempts = 0
    max_link_verification_attempts = 3
    
    for attempt in range(max_retries):
        try:
            response = Generation.call(
                model=model,
                prompt=prompt,
                temperature=temperature,
                top_p=top_p,
                max_tokens=max_tokens
            )
            if response.status_code == 200 and response.output and response.output.text:
                raw_response = response.output.text.strip()
                
                # 尝试解析并验证链接
                policy_name, link = parse_llm_response(raw_response)
                
                if policy_name and link:
                    # ✅ 成功获得有效链接
                    return raw_response
                else:
                    # 链接无效，记录但继续重试
                    valid_link_attempts += 1
                    if valid_link_attempts >= max_link_verification_attempts:
                        # 已重试多次都没有有效链接，返回空
                        return ""
                    # 否则继续下一个 API 调用
                    continue
                    
        except Exception as e:
            pass
    
    return ""

def append_to_excel(output_path, record):
    """将单条记录追加到 Excel 文件末尾（兼容首次创建，支持去重）"""
    # 检查是否已存在相同的记录（基于政策名称、国家和议题）
    if os.path.exists(output_path):
        try:
            existing_df = pd.read_excel(output_path)
            # 检查是否已存在相同的记录
            duplicate_mask = (
                (existing_df["policy_name"] == record["policy_name"]) &
                (existing_df["country_en"] == record["country_en"]) &
                (existing_df["link"] == record["link"])  # 也检查链接是否相同
            )
            if duplicate_mask.any():
                print(f"  ⚠️ 跳过重复记录: {record['policy_name'][:30]}...")
                return False  # 返回 False 表示未保存（重复）
        except Exception as e:
            print(f"  ⚠️ 读取现有文件失败，将创建新文件: {e}")

    # 如果不存在重复记录，则追加
    df = pd.DataFrame([record])
    if os.path.exists(output_path):
        # 获取当前最大行号
        book = load_workbook(output_path)
        sheet = book.active
        startrow = sheet.max_row
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=startrow)
    else:
        # 首次创建，写入带表头的文件
        df.to_excel(output_path, index=False)

    return True  # 返回 True 表示成功保存

# ==============================
# 主逻辑
# ==============================
def get_country_selection(countries):
    """通过序号选择要处理的国家"""
    print("\n📍 国家选择")
    print("="*50)
    print(f"✅ 数据库中共有 {len(countries)} 个国家")
    print("\n")

    country_labels = [f"{cn} ({en})" for cn, en, _ in countries]
    selected_labels = select_items(country_labels, "国家(country)", allow_multiple=True)

    if not selected_labels:
        # 如果用户取消或未选择，则默认全部国家
        print("⚠️ 未选择有效国家，默认处理全部国家")
        return countries

    label_to_country = {f"{cn} ({en})": (cn, en, iso) for cn, en, iso in countries}
    selected_countries = [label_to_country[label] for label in selected_labels if label in label_to_country]

    print(f"✅ 已选择 {len(selected_countries)} 个国家")
    return selected_countries

def select_items(items, item_type, allow_multiple=True):
    """通用选择函数，支持单选或多选"""
    print(f"\n📋 {item_type}选择")
    print("="*50)
    print(f"✅ 数据库中共有 {len(items)} 个{item_type}")
    print("\n可选项目：")
    
    for i, item in enumerate(items, 1):
        print(f"{i:2d}. {item}")
    
    print("\n")
    
    if allow_multiple:
        print("💡 输入格式：")
        print("   - 单个序号：5")
        print("   - 多个序号：1,3,5 或 1-3,7")
        print("   - 全部选择：all 或回车")
        print("   - 随机选择：random N（选择N个随机项目）")
    else:
        print("💡 输入格式：")
        print("   - 序号：5")
        print("   - 全部：all 或回车")
    
    while True:
        try:
            choice = input(f"请选择{item_type}：").strip().lower()
            
            if not choice or choice == "all":
                return items
            
            if allow_multiple and choice.startswith("random "):
                try:
                    count = int(choice.split()[1])
                    if count <= 0 or count > len(items):
                        print(f"❌ 随机数量必须在 1-{len(items)} 之间")
                        continue
                    import random
                    selected = random.sample(items, count)
                    print(f"🎲 随机选择了 {count} 个{item_type}：")
                    for item in selected:
                        print(f"   • {item}")
                    return selected
                except (ValueError, IndexError):
                    print("❌ 随机选择格式错误，请使用 'random N' 格式")
                    continue
            
            selected_items = []
            
            # 解析输入（如 "1,3,5" 或 "1-3,7"）
            parts = choice.replace(" ", "").split(",")
            for part in parts:
                if "-" in part:
                    # 处理范围（如 "1-3"）
                    start_end = part.split("-")
                    if len(start_end) == 2:
                        try:
                            start = int(start_end[0]) - 1  # 转换为0-based索引
                            end = int(start_end[1])  # 转换为1-based索引
                            if 0 <= start < end <= len(items):
                                selected_items.extend(items[start:end])
                            else:
                                print(f"❌ 范围 {part} 超出有效范围")
                                selected_items = []
                                break
                        except ValueError:
                            print(f"❌ 范围格式错误: {part}")
                            selected_items = []
                            break
                    else:
                        print(f"❌ 范围格式错误: {part}")
                        selected_items = []
                        break
                else:
                    # 处理单个序号
                    try:
                        idx = int(part) - 1  # 转换为0-based索引
                        if 0 <= idx < len(items):
                            selected_items.append(items[idx])
                        else:
                            print(f"❌ 序号 {int(part)} 超出范围 (1-{len(items)})")
                            selected_items = []
                            break
                    except ValueError:
                        print(f"❌ 无效序号: {part}")
                        selected_items = []
                        break
            
            if selected_items:
                # 去重并保持原始顺序
                seen = set()
                unique_selected = []
                for item in selected_items:
                    if item not in seen:
                        seen.add(item)
                        unique_selected.append(item)
                
                print(f"✅ 已选择 {len(unique_selected)} 个{item_type}：")
                for item in unique_selected:
                    print(f"   • {item}")
                return unique_selected
            else:
                print("❌ 没有有效的选择，请重新输入")
                
        except KeyboardInterrupt:
            print("\n❌ 操作已取消")
            return []
        except Exception as e:
            print(f"❌ 输入错误: {e}")

def get_topic_selection(topics):
    """获取用户选择的议题"""
    return select_items(topics, "ESG议题(topic)", allow_multiple=True)

def get_processing_mode():
    """获取用户选择的数据处理模式"""
    print("\n📝 数据处理模式")
    print("="*50)
    print("选择如何处理现有数据：")
    print("1. 追加模式 (默认) - 在现有数据基础上追加新记录，自动跳过重复")
    print("2. 重置模式 - 清空现有文件，重新开始收集")
    print("\n")
    
    while True:
        choice = input("请选择模式 (1-2，默认1): ").strip()
        if choice == "2":
            # 确认是否真的要清空
            confirm = input("⚠️ 确定要清空所有现有数据吗？(输入 'yes' 确认): ").strip().lower()
            if confirm == "yes":
                return "reset"
            else:
                print("已取消重置操作，使用追加模式")
                return "append"
        elif choice in ["", "1"]:
            return "append"
        else:
            print("❌ 输入有误！请输入 1 或 2")

def generate_policy_links_with_llm():
    all_countries, all_topics = load_countries_and_topics()
    today = datetime.now().strftime("%Y-%m-%d")
    
    # 通过序号选择国家
    countries = get_country_selection(all_countries)
    
    # 获取用户选择的议题
    selected_topics = get_topic_selection(all_topics)
    
    # 获取数据处理模式
    processing_mode = get_processing_mode()
    
    # 如果选择重置模式，清空现有文件
    if processing_mode == "reset" and os.path.exists(OUTPUT_PATH):
        try:
            os.remove(OUTPUT_PATH)
            print(f"🗑️ 已清空现有文件: {OUTPUT_PATH}")
        except Exception as e:
            print(f"⚠️ 清空文件失败: {e}")
    
    # 获取大模型配置
    model_config = select_llm_model()
    
    print("\n🚀 开始逐国家、逐议题检索政策链接（确保所有链接有效），并实时保存...")
    print(f"📊 处理进度: 0/{len(countries)} 国家, {len(selected_topics)} 议题")
    print(f"💡 总组合数: {len(countries)} × {len(selected_topics)} = {len(countries) * len(selected_topics)}")
    print()
    
    total_saved = 0
    total_skipped = 0
    total_invalid = 0
    
    for idx, (country_cn, country_en, iso_alpha3) in enumerate(countries, 1):
        print(f"\n[{idx}/{len(countries)}] 📍 国家: {country_cn} ({country_en})")
        
        for topic in selected_topics:
            prompt = generate_policy_prompt(country_en, topic)
            response = call_qwen(prompt, model_config, country_en, topic)
            
            policy_name, link = parse_llm_response(response)
            if policy_name and link:
                policy_id = "P" + uuid.uuid4().hex[:6].upper()
                record = {
                    "policy_id": policy_id,
                    "policy_name": policy_name,
                    "country_cn": country_cn,
                    "country_en": country_en,
                    "iso_alpha3": iso_alpha3,
                    "topic": topic,    # 保留议题信息
                    "link": link,
                    "record_date": today
                }
                # ✅ 关键：立即保存，不等待全部完成，支持去重
                if append_to_excel(OUTPUT_PATH, record):
                    total_saved += 1
                    print(f"    ✅ [{topic}] {policy_name[:35]}... → 已保存 (链接有效)")
                else:
                    total_skipped += 1
                    print(f"    ⏭️ [{topic}] {policy_name[:35]}... → 已跳过 (重复记录)")
            else:
                total_invalid += 1
                print(f"    ❌ [{topic}] 未找到有效政策链接或链接无法访问")

    print(f"\n{'='*60}")
    print(f"🎉 检索完成！")
    print(f"📊 已处理: {len(countries)} 个国家, {len(selected_topics)} 个议题")
    print(f"💾 新增保存: {total_saved} 条记录")
    print(f"⏭️ 跳过重复: {total_skipped} 条记录")
    print(f"⚠️  未保存: {total_invalid} 条（无有效链接或链接无法访问）")
    print(f"📁 输出文件: {OUTPUT_PATH}")
    print(f"{'='*60}")

# ==============================
# 入口
# ==============================
if __name__ == "__main__":
    generate_policy_links_with_llm()